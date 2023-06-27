# Selenium模块是一个自动化测试工具，能够驱动浏览器模拟人的操作，如单击、键盘输入等。
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
import re
import pandas as pd
import time
import jieba
from jieba.analyse import *
from openpyxl import load_workbook
from ciyuntu import WordCloud
import matplotlib.pyplot


txt=[]

# 从获取的网页源代码中提取目标数据
def extract_data(html_code):
    # 目标数据的正则表达式
    p_time='<span class="tb-r-date">(.*?)</span>'
    p_comment='<div class="J_KgRate_ReviewContent tb-tbcr-content ">(.*?)</div>'

    # 利用findall()函数提取目标数据
    time=re.findall(p_time, html_code, re.S)
    comment=re.findall(p_comment,html_code,re.S)

    # 将几个目标数据列表转换为一个字典
    data_dt = {'评论时间': time, '评论': comment}
    # 用上面的字典创建一个DataFrame
    return pd.DataFrame(data_dt)


def get_pages(start, end):
    # 声明要模拟的浏览器是Chrome,并启用无界面浏览模式
    # chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    # browser = webdriver.Chrome(options=chrome_options)
    browser = webdriver.Chrome()
    browser.maximize_window()

    # 通过get()函数控制浏览器发起请求，访问网址,获取源码
    #TODO:
    url = 'https://item.taobao.com/item.htm?id=716719530600&ali_refid=a3_430673_1006:1104967086:N:2CMPy9e41kyMVNEdUOKF6Q%3D%3D:fbfae5fa9eed1d0bae53032c341f85ed&ali_trackid=1_fbfae5fa9eed1d0bae53032c341f85ed&spm=a2e0b.20350158.31919782.1'
    browser.get(url)

    #如要验证，在20秒内扫码登录即可
    time.sleep(20)

    # 模拟人操作浏览器，输入搜索关键词，点击搜索按钮
    #评论按钮
    browser.find_element(By.XPATH, '//*[@id="J_TabBar"]/li[2]/a').click()
    time.sleep(random.randint(5, 10))

    all_data = pd.DataFrame()

    for page in range(start, end + 1):
        # 模拟人操作浏览器，输入搜索关键词，点击搜索按钮
        # browser.find_element(By.XPATH, '//*[@id="jump_page"]').clear()
        # browser.find_element(By.XPATH, '//*[@id="jump_page"]').send_keys(page)
        browser.find_element(By.XPATH, '//*[@id="reviews"]/div[2]/div/div/div/div/div[2]/div/ul/li[5]').click()
        # 等待浏览器与服务器交互刷新数据，否则获取不到动态信息
        time.sleep(random.randint(5, 10))
        # 将提取的目标数据添加到DataFrame中
        all_data = all_data._append(extract_data(browser.page_source))

    browser.quit()

    # 将DataFrame保存为Excel
    all_data.to_excel('评价.xlsx', index=False)


def word_cloud():
    #读取xlsx
    workbook=load_workbook('评价.xlsx')
    worksheet=workbook['Sheet1']
    data=[]
    f=open('评论信息.txt','w')
    for row in range(2,worksheet.max_row+1):
        data.append(worksheet['B'+str(row)].value)
    f.write(str(data))
    f.close()

    # #去html标签
    # f = open('评价信息.txt','r')
    # g = re.sub('<.*?>','',f)
    # g = re.sub('nbsp', '', g)







#运行参数分别为评论的起始页数和结束页数
get_pages(1, 3)

