import random
from selenium import webdriver
from selenium.webdriver.common.by import By
import re
import pandas as pd
import time
import jieba
from jieba.analyse import *
from openpyxl import load_workbook
from wordcloud import WordCloud
import matplotlib.pyplot as plt

def word_cloud():
    #读取xlsx
    workbook=load_workbook('评价.xlsx')
    worksheet=workbook['Sheet1']
    data = []
    f = open('评论信息.txt', 'w')
    for row in range(2, worksheet.max_row + 1):
        data.append(worksheet['B'+str(row)].value)

    f.write(str(data))
    f.close()

    #去html标签
    f = open('评论信息.txt', 'r').read()
    g = re.sub("\[\'", '', f)
    g = re.sub("\'\]", '', g)
    g = re.sub("\', \'", '', g)
    g = re.sub(" ", '', g)

    h = open('评论信息-清洗后.txt','w')
    h.write(g)
    h.close()

    f=open('评论信息-清洗后.txt', 'r', encoding='gbk').read()
    sep_list=jieba.lcut(f)

    stopwords = [line.strip() for line in open('stop.txt', 'r', encoding='utf-8').readlines()]
    outstr = ''
    for word in sep_list:
        if word not in stopwords:
            outstr += word

    outstr = jieba.lcut(outstr)
    outstr = ' '.join(outstr)

    # 生成词云图
    # 设置词云使用的字体
    font = r'C:\Windows\Fonts\simsun.ttc'
    wc = WordCloud(font_path=font, width=2400, height=1200, max_words=100)
    wc.generate(outstr)
    wc.to_file('词云.jpg')
    plt.figure(dpi=100)
    plt.imshow(wc, interpolation='catrom')
    plt.axis('off')
    plt.show()
    plt.close()

word_cloud()
